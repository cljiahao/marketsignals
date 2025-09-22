import os
from re import S
from openpyxl import *

class Database:
    def __init__(self,array,country):
        self.array = array
        self.country = country
        self.main()

    def main(self):
        
        dataPath = os.path.join(os.path.dirname(__file__),"Database")
        if not os.path.exists(dataPath): os.makedirs(dataPath)

        filePath = os.path.join(dataPath,"Database.xlsx")

        if os.path.exists(filePath):
            
            wb = load_workbook(os.path.join(dataPath,"Database.xlsx"))

            if self.country in wb.sheetnames: 
                ws = wb[self.country] 
            else:
                ws = wb.create_sheet(self.country)
                header = ["Date","Ticker","TimeFrame","Buy","Sell","TimeFrame","Buy","Sell"]
                for i in range(len(header)):
                    ws.cell(row=1,column=i+1).value = header[i]

            maxrow = ws.max_row
            for j in range(len(self.array)):
                ws.cell(row=maxrow+1,column=j+1).value = self.array[j]

            wb.save(filePath)

        else:

            wb = Workbook()
            ws = wb.create_sheet(self.country)
            header = ["Date","Ticker","TimeFrame","Buy","Sell","TimeFrame","Buy","Sell"]
            for i in range(len(header)):
                ws.cell(row=1,column=i+1).value = header[i]

            maxrow = ws.max_row
            for j in range(len(self.array)):
                ws.cell(row=maxrow+1,column=j+1).value = self.array[j]

            wb.save(filePath)

